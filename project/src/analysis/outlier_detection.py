"""
Outlier Detection for Taninaka (谷中健介) Data

このスクリプトは、taninaka（谷中健介）のデータが統計的に外れ値であるかを検証します。

検証方法:
1. Grubbs検定（最大値/最小値の外れ値検定）
2. IQR法（四分位範囲法）
3. z-scoreベースの検定
4. 可視化（箱ひげ図、分布図）

データセット:
- Performance (タスク遂行時間)
- Subjective (主観評価)
- NASA-TLX (認知負荷)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Tuple, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 日本語フォント設定
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'Hiragino Sans', 'Yu Gothic', 'Meirio', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

class OutlierDetector:
    """外れ値検出クラス"""

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.results = []

    def grubbs_test(self, data: np.ndarray, alpha: float = 0.05) -> Tuple[float, float, bool, str]:
        """
        Grubbs検定（最大値または最小値が外れ値かを検定）

        Returns:
            G統計量, 臨界値, 外れ値判定, 外れ値の種類（max/min）
        """
        n = len(data)
        mean = np.mean(data)
        std = np.std(data, ddof=1)

        # 最大値と最小値のG統計量を計算
        abs_diff = np.abs(data - mean)
        G_max_idx = np.argmax(abs_diff)
        G_max = abs_diff[G_max_idx] / std

        # 臨界値を計算（両側検定）
        t_dist = stats.t.ppf(1 - alpha / (2 * n), n - 2)
        G_critical = ((n - 1) / np.sqrt(n)) * np.sqrt(t_dist**2 / (n - 2 + t_dist**2))

        is_outlier = G_max > G_critical
        outlier_type = 'max' if data[G_max_idx] == np.max(data) else 'min'

        return G_max, G_critical, is_outlier, outlier_type

    def iqr_method(self, data: np.ndarray, k: float = 1.5) -> Tuple[List[int], float, float]:
        """
        IQR法による外れ値検出

        Returns:
            外れ値のインデックス, 下限, 上限
        """
        Q1 = np.percentile(data, 25)
        Q3 = np.percentile(data, 75)
        IQR = Q3 - Q1

        lower_bound = Q1 - k * IQR
        upper_bound = Q3 + k * IQR

        outliers_idx = np.where((data < lower_bound) | (data > upper_bound))[0]

        return list(outliers_idx), lower_bound, upper_bound

    def zscore_method(self, data: np.ndarray, threshold: float = 3.0) -> List[int]:
        """
        z-scoreによる外れ値検出

        Returns:
            外れ値のインデックス
        """
        z_scores = np.abs(stats.zscore(data, ddof=1))
        outliers_idx = np.where(z_scores > threshold)[0]
        return list(outliers_idx)

    def detect_outliers(self, data: pd.Series, participant_names: List[str],
                       dataset_name: str, metric_name: str) -> Dict:
        """
        複数の方法で外れ値を検出

        Args:
            data: 各参加者の値（平均値など）
            participant_names: 参加者名のリスト
            dataset_name: データセット名
            metric_name: 指標名
        """
        data_array = data.values

        # 1. Grubbs検定
        G_stat, G_crit, is_outlier_grubbs, outlier_type = self.grubbs_test(data_array)

        # 2. IQR法
        iqr_outliers, iqr_lower, iqr_upper = self.iqr_method(data_array)

        # 3. z-score法
        zscore_outliers = self.zscore_method(data_array)

        # taninakaのインデックスを探す
        taninaka_idx = None
        for i, name in enumerate(participant_names):
            if 'taninaka' in name.lower() or '谷中' in name:
                taninaka_idx = i
                break

        result = {
            'dataset': dataset_name,
            'metric': metric_name,
            'n': len(data_array),
            'mean': np.mean(data_array),
            'std': np.std(data_array, ddof=1),
            'median': np.median(data_array),
            'min': np.min(data_array),
            'max': np.max(data_array),
            'participants': participant_names,
            'values': data_array,
            'taninaka_idx': taninaka_idx,
            'taninaka_value': data_array[taninaka_idx] if taninaka_idx is not None else None,
            'grubbs': {
                'G_stat': G_stat,
                'G_critical': G_crit,
                'is_outlier': is_outlier_grubbs,
                'outlier_type': outlier_type,
                'p_value': 0.05  # 使用したalpha
            },
            'iqr': {
                'outliers_idx': iqr_outliers,
                'lower_bound': iqr_lower,
                'upper_bound': iqr_upper,
                'is_taninaka_outlier': taninaka_idx in iqr_outliers if taninaka_idx is not None else False
            },
            'zscore': {
                'outliers_idx': zscore_outliers,
                'threshold': 3.0,
                'is_taninaka_outlier': taninaka_idx in zscore_outliers if taninaka_idx is not None else False
            }
        }

        self.results.append(result)
        return result

    def visualize_outliers(self, result: Dict, save_path: Path):
        """外れ値検出結果の可視化"""
        fig, axes = plt.subplots(1, 2, figsize=(14, 6))

        data = result['values']
        participants = result['participants']
        taninaka_idx = result['taninaka_idx']

        # 箱ひげ図
        ax1 = axes[0]
        bp = ax1.boxplot([data], vert=True, patch_artist=True, widths=0.5)
        bp['boxes'][0].set_facecolor('lightblue')

        # 個別データポイントをプロット
        x_pos = np.ones(len(data))
        colors = ['red' if i == taninaka_idx else 'blue' for i in range(len(data))]
        sizes = [100 if i == taninaka_idx else 50 for i in range(len(data))]

        for i, (x, y) in enumerate(zip(x_pos, data)):
            ax1.scatter(x, y, c=colors[i], s=sizes[i], alpha=0.6, zorder=3,
                       label='Taninaka' if i == taninaka_idx else '')

        # IQR境界線
        ax1.axhline(result['iqr']['lower_bound'], color='orange', linestyle='--',
                   linewidth=1.5, label=f"IQR Lower: {result['iqr']['lower_bound']:.2f}")
        ax1.axhline(result['iqr']['upper_bound'], color='orange', linestyle='--',
                   linewidth=1.5, label=f"IQR Upper: {result['iqr']['upper_bound']:.2f}")

        ax1.set_ylabel('Value', fontsize=12)
        ax1.set_title(f"{result['dataset']} - {result['metric']}\nBoxplot with IQR bounds",
                     fontsize=14, fontweight='bold')
        ax1.legend(loc='best', fontsize=10)
        ax1.grid(True, alpha=0.3)

        # 棒グラフ（参加者別）
        ax2 = axes[1]
        x_positions = np.arange(len(data))
        bar_colors = ['red' if i == taninaka_idx else 'steelblue' for i in range(len(data))]

        bars = ax2.bar(x_positions, data, color=bar_colors, alpha=0.7, edgecolor='black')

        # 平均線
        mean_val = result['mean']
        ax2.axhline(mean_val, color='green', linestyle='-', linewidth=2,
                   label=f'Mean: {mean_val:.2f}')

        # ±2SD線
        std_val = result['std']
        ax2.axhline(mean_val + 2*std_val, color='purple', linestyle='--',
                   linewidth=1.5, label=f'+2SD: {mean_val + 2*std_val:.2f}')
        ax2.axhline(mean_val - 2*std_val, color='purple', linestyle='--',
                   linewidth=1.5, label=f'-2SD: {mean_val - 2*std_val:.2f}')

        ax2.set_xlabel('Participant', fontsize=12)
        ax2.set_ylabel('Value', fontsize=12)
        ax2.set_title(f"{result['dataset']} - {result['metric']}\nParticipant-wise Distribution",
                     fontsize=14, fontweight='bold')
        ax2.set_xticks(x_positions)
        ax2.set_xticklabels([p[:10] for p in participants], rotation=45, ha='right', fontsize=9)
        ax2.legend(loc='best', fontsize=10)
        ax2.grid(True, alpha=0.3, axis='y')

        plt.tight_layout()
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()

        print(f"  ✓ 可視化保存: {save_path.name}")

    def print_result(self, result: Dict):
        """結果を表示"""
        print(f"\n{'='*80}")
        print(f"データセット: {result['dataset']} - {result['metric']}")
        print(f"{'='*80}")
        print(f"サンプルサイズ: n = {result['n']}")
        print(f"平均: {result['mean']:.4f}")
        print(f"標準偏差: {result['std']:.4f}")
        print(f"中央値: {result['median']:.4f}")
        print(f"範囲: [{result['min']:.4f}, {result['max']:.4f}]")

        if result['taninaka_idx'] is not None:
            print(f"\nTaninakaの値: {result['taninaka_value']:.4f}")
            print(f"平均との差: {result['taninaka_value'] - result['mean']:.4f} "
                  f"({(result['taninaka_value'] - result['mean']) / result['std']:.2f} SD)")

        print(f"\n【Grubbs検定】")
        print(f"  G統計量: {result['grubbs']['G_stat']:.4f}")
        print(f"  臨界値 (α=0.05): {result['grubbs']['G_critical']:.4f}")
        print(f"  判定: {'✗ 外れ値あり' if result['grubbs']['is_outlier'] else '✓ 外れ値なし'}")
        if result['grubbs']['is_outlier']:
            print(f"  外れ値の種類: {result['grubbs']['outlier_type']}")

        print(f"\n【IQR法 (k=1.5)】")
        print(f"  下限: {result['iqr']['lower_bound']:.4f}")
        print(f"  上限: {result['iqr']['upper_bound']:.4f}")
        print(f"  外れ値数: {len(result['iqr']['outliers_idx'])}")
        if result['taninaka_idx'] is not None:
            print(f"  Taninaka判定: {'✗ 外れ値' if result['iqr']['is_taninaka_outlier'] else '✓ 正常範囲'}")

        print(f"\n【z-score法 (threshold=3.0)】")
        print(f"  外れ値数: {len(result['zscore']['outliers_idx'])}")
        if result['taninaka_idx'] is not None:
            print(f"  Taninaka判定: {'✗ 外れ値' if result['zscore']['is_taninaka_outlier'] else '✓ 正常範囲'}")

    def save_results_to_excel(self, output_path: Path):
        """結果をExcelに保存"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        # サマリーシート
        ws_summary = wb.create_sheet("Summary")
        headers = ['Dataset', 'Metric', 'n', 'Mean', 'SD', 'Median', 'Min', 'Max',
                  'Taninaka Value', 'Taninaka vs Mean (SD)',
                  'Grubbs: Outlier?', 'Grubbs: G-stat', 'Grubbs: G-crit',
                  'IQR: Taninaka Outlier?', 'IQR: Lower', 'IQR: Upper',
                  'Z-score: Taninaka Outlier?']

        ws_summary.append(headers)

        # ヘッダーのスタイル
        for cell in ws_summary[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for result in self.results:
            taninaka_vs_mean = (result['taninaka_value'] - result['mean']) / result['std'] if result['taninaka_value'] is not None else None

            row = [
                result['dataset'],
                result['metric'],
                result['n'],
                result['mean'],
                result['std'],
                result['median'],
                result['min'],
                result['max'],
                result['taninaka_value'] if result['taninaka_value'] is not None else 'N/A',
                taninaka_vs_mean if taninaka_vs_mean is not None else 'N/A',
                'YES' if result['grubbs']['is_outlier'] else 'NO',
                result['grubbs']['G_stat'],
                result['grubbs']['G_critical'],
                'YES' if result['iqr']['is_taninaka_outlier'] else 'NO',
                result['iqr']['lower_bound'],
                result['iqr']['upper_bound'],
                'YES' if result['zscore']['is_taninaka_outlier'] else 'NO'
            ]
            ws_summary.append(row)

        # 列幅の調整
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_summary.column_dimensions[column_letter].width = adjusted_width

        # 詳細データシート
        for result in self.results:
            sheet_name = f"{result['dataset']}_{result['metric']}"[:31]  # Excel制限
            ws = wb.create_sheet(sheet_name)

            ws.append(['Participant', 'Value', 'Is Taninaka?'])
            for i, (name, val) in enumerate(zip(result['participants'], result['values'])):
                is_taninaka = 'YES' if i == result['taninaka_idx'] else 'NO'
                ws.append([name, val, is_taninaka])

        wb.save(output_path)
        print(f"\n✓ Excel保存: {output_path}")


def analyze_performance_data(detector: OutlierDetector):
    """Performanceデータの外れ値検証"""
    print("\n" + "="*80)
    print("PERFORMANCE データの外れ値検証")
    print("="*80)

    csv_path = Path('/Users/toshiki/Desktop/dev/comic-effect/project/files/performance.csv')
    df = pd.read_csv(csv_path)

    # 参加者ごとの全条件平均を計算
    participant_col = 'name'
    task_col = 'task'
    time_cols = ['no1', 'no2', 'no3']
    condition_cols = ['a', 'b', 'c', 'd', 'e']

    # 各参加者の平均を計算（3試行の平均値を5条件で平均）
    participant_means = {}

    for participant in df[participant_col].unique():
        participant_data = df[df[participant_col] == participant]
        condition_means = []

        for cond in condition_cols:
            cond_data = participant_data[participant_data[task_col] == cond]
            if len(cond_data) > 0:
                # 各条件の3試行の平均
                condition_means.append(cond_data[time_cols].mean().mean())

        # 5条件の平均
        if len(condition_means) > 0:
            participant_means[participant] = np.mean(condition_means)

    participants = list(participant_means.keys())
    values = pd.Series(list(participant_means.values()))

    result = detector.detect_outliers(values, participants, 'Performance', 'Mean Task Time')
    detector.print_result(result)

    # 可視化
    viz_path = detector.output_dir / 'performance_outlier_detection.png'
    detector.visualize_outliers(result, viz_path)

    return result


def analyze_subjective_data(detector: OutlierDetector):
    """Subjectiveデータの外れ値検証"""
    print("\n" + "="*80)
    print("SUBJECTIVE データの外れ値検証")
    print("="*80)

    csv_path = Path('/Users/toshiki/Desktop/dev/comic-effect/project/files/subjective.csv')
    df = pd.read_csv(csv_path)

    # 参加者ごとの全項目・全条件の平均を計算
    participant_col = '氏名'
    condition_col = '実験タイプを教えてください'
    # 評価項目は4列目から9列目まで（6項目）
    item_cols = df.columns[4:10]

    participant_means = {}

    for participant in df[participant_col].unique():
        participant_data = df[df[participant_col] == participant]
        # 数値に変換して平均を計算
        all_values = participant_data[item_cols].apply(pd.to_numeric, errors='coerce').values.flatten()
        # NaNを除外
        all_values = all_values[~np.isnan(all_values)]
        if len(all_values) > 0:
            participant_means[participant] = np.mean(all_values)

    participants = list(participant_means.keys())
    values = pd.Series(list(participant_means.values()))

    result = detector.detect_outliers(values, participants, 'Subjective', 'Mean Rating (All Items)')
    detector.print_result(result)

    # 可視化
    viz_path = detector.output_dir / 'subjective_outlier_detection.png'
    detector.visualize_outliers(result, viz_path)

    # Readabilityのみでも検証
    readability_col = df.columns[4]  # 最初の評価項目（読みやすさ）
    participant_means_readability = {}

    for participant in df[participant_col].unique():
        participant_data = df[df[participant_col] == participant]
        # 数値に変換
        readability_values = pd.to_numeric(participant_data[readability_col], errors='coerce')
        participant_means_readability[participant] = readability_values.mean()

    participants_r = list(participant_means_readability.keys())
    values_r = pd.Series(list(participant_means_readability.values()))

    result_r = detector.detect_outliers(values_r, participants_r, 'Subjective', 'Readability Only')
    detector.print_result(result_r)

    # 可視化
    viz_path_r = detector.output_dir / 'subjective_readability_outlier_detection.png'
    detector.visualize_outliers(result_r, viz_path_r)

    return result, result_r


def analyze_nasatlx_data(detector: OutlierDetector):
    """NASA-TLXデータの外れ値検証"""
    print("\n" + "="*80)
    print("NASA-TLX データの外れ値検証")
    print("="*80)

    nasatlx_dir = Path('/Users/toshiki/Desktop/dev/comic-effect/project/files/nasa-tlx')

    # 参加者ごとの全条件平均を計算
    participant_means = {}

    for csv_file in sorted(nasatlx_dir.glob('*.csv')):
        # ファイル名から参加者名と条件を抽出
        filename = csv_file.stem
        parts = filename.split('_')

        if len(parts) >= 2:
            participant = parts[0]
            condition = parts[1]

            # CSVを読み込んで総スコアを取得
            df_csv = pd.read_csv(csv_file)
            if len(df_csv) >= 7:
                total_score = df_csv.iloc[-1, -1]  # 最後の行、最後の列
                try:
                    total_score = float(total_score)

                    if participant not in participant_means:
                        participant_means[participant] = []
                    participant_means[participant].append(total_score)
                except (ValueError, TypeError):
                    continue

    # 各参加者の5条件平均を計算
    participant_avg = {p: np.mean(scores) for p, scores in participant_means.items() if len(scores) > 0}

    participants = list(participant_avg.keys())
    values = pd.Series(list(participant_avg.values()))

    result = detector.detect_outliers(values, participants, 'NASA-TLX', 'Mean Workload Score')
    detector.print_result(result)

    # 可視化
    viz_path = detector.output_dir / 'nasatlx_outlier_detection.png'
    detector.visualize_outliers(result, viz_path)

    return result


def main():
    """メイン処理"""
    print("\n" + "="*80)
    print("外れ値検証スクリプト - Taninaka（谷中健介）データの統計的検証")
    print("="*80)
    print("\n検証方法:")
    print("  1. Grubbs検定（最大/最小値の外れ値検定）")
    print("  2. IQR法（四分位範囲法、k=1.5）")
    print("  3. z-score法（threshold=3.0）")
    print("="*80)

    # 出力ディレクトリ
    output_dir = Path('/Users/toshiki/Desktop/dev/comic-effect/project/files/outlier_detection')
    detector = OutlierDetector(output_dir)

    # 各データセットで検証
    perf_result = analyze_performance_data(detector)
    subj_result, subj_read_result = analyze_subjective_data(detector)
    nasa_result = analyze_nasatlx_data(detector)

    # 結果をExcelに保存
    excel_path = output_dir / 'outlier_detection_results.xlsx'
    detector.save_results_to_excel(excel_path)

    # 総合まとめ
    print("\n" + "="*80)
    print("総合まとめ")
    print("="*80)

    all_results = [perf_result, subj_result, subj_read_result, nasa_result]

    taninaka_outlier_count = 0
    total_tests = 0

    for result in all_results:
        if result['taninaka_idx'] is not None:
            total_tests += 3  # Grubbs, IQR, z-score

            # 外れ値と判定された数をカウント
            if result['grubbs']['is_outlier']:
                # Grubbsで検出された外れ値がtaninakaか確認
                data = result['values']
                taninaka_val = result['taninaka_value']
                if result['grubbs']['outlier_type'] == 'max' and taninaka_val == np.max(data):
                    taninaka_outlier_count += 1
                elif result['grubbs']['outlier_type'] == 'min' and taninaka_val == np.min(data):
                    taninaka_outlier_count += 1

            if result['iqr']['is_taninaka_outlier']:
                taninaka_outlier_count += 1

            if result['zscore']['is_taninaka_outlier']:
                taninaka_outlier_count += 1

    print(f"\nTaninakaが外れ値と判定された回数: {taninaka_outlier_count} / {total_tests}")
    print(f"外れ値判定率: {taninaka_outlier_count / total_tests * 100:.1f}%")

    print("\n結論:")
    if taninaka_outlier_count / total_tests >= 0.5:
        print("  ✓ Taninakaのデータは統計的に外れ値と判定される傾向が強い")
        print("  → データから除外することは統計的に正当化される")
    else:
        print("  ⚠ Taninakaのデータが外れ値と判定されるケースは限定的")
        print("  → 除外の妥当性について慎重な検討が必要")

    print(f"\n✓ すべての結果を保存しました: {output_dir}")
    print(f"  - Excel: {excel_path.name}")
    print(f"  - 可視化: *.png")


if __name__ == '__main__':
    main()
